"""差异台账 Excel 导出器"""
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ExcelExporter:
    """Excel台账导出器"""

    def __init__(self):
        self.default_headers = [
            '序号', '状态', '归因类型', '归因原因',
            'HIS记录', '医保记录', '差异字段', '差异值(HIS)', '差异值(医保)'
        ]

    def export(self, diff_records: List[Dict[str, Any]], output_path: str,
               filter_type: Optional[str] = None) -> None:
        """导出差异台账到Excel

        Args:
            diff_records: 差异记录列表
            output_path: 输出文件路径
            filter_type: 按归因类型筛选（可选）
        """
        # 按类型筛选
        if filter_type:
            diff_records = [r for r in diff_records if r.get('attribution_type') == filter_type]

        wb = Workbook()
        ws = wb.active
        ws.title = "差异台账"

        # 写入表头
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(self.default_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 写入数据
        for row_idx, record in enumerate(diff_records, 2):
            his_record = record.get('his_record', {})
            ins_record = record.get('insurance_record', {})
            diff_values = record.get('diff_values', {})

            # 序列化成字符串
            his_str = str(his_record) if his_record else ''
            ins_str = str(ins_record) if ins_record else ''

            row_data = [
                row_idx - 1,  # 序号
                record.get('status', ''),
                record.get('attribution_type', ''),
                record.get('attribution_reason', ''),
                his_str[:200] if len(his_str) > 200 else his_str,  # 截断过长内容
                ins_str[:200] if len(ins_str) > 200 else ins_str,
                ', '.join(record.get('diff_fields', [])),
                str(list(diff_values.keys())),
                str(diff_values)
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 30

        # 冻结首行
        ws.freeze_panes = 'A2'

        wb.save(output_path)
